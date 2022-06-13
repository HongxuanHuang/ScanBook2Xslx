"""
@ Author: GVenusLeo
@ Email: gvenusleo@gmail.com
@ Website: https://promiseland.top
@ Update: 2022-03-15
@ Description: 扫码获取图书信息并写入Excel表格
"""

from pywebio import pin, start_server, config
from pywebio.output import put_html, put_row, put_column, use_scope, put_scope, put_buttons, toast
from config import *
from requests import get
from openpyxl import Workbook, load_workbook
import json


global book_info
global num
global isbn


def get_book_info(isbn):
    """
    获取图书信息
    :param isbn: 图书条形码
    :return: 图书信息
    """
    url = "https://api.jike.xyz/situ/book/isbn/{}?apikey={}".format(isbn, APIKEY)
    response = get(url)
    response = json.loads(response.text)
    if response["msg"] != "请求成功" or response["data"] is None:
        return "Error"
    else:
        return response["data"]


def in_sheet():
    """
    图书数据写入表格book.xslx
    :return: None
    """
    global num
    id = pin.pin["id"]
    if id == "":
        toast(
            content="请输入图书ID！",
            duration=2,
            position="center",
            color="error"
        )
    elif len(isbn) != 13:
        toast(
            content="请输入正确的图书ISBN！",
            duration=2,
            position="center",
            color="error"
        )
        pin.pin["isbn"] = ""
    else:
        name = book_info["name"]
        author = book_info["author"]
        publisher = book_info["publishing"]
        try:
            wb = load_workbook(filename="book.xlsx")
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "书名", "出版社", "作者"])
        ws.append([id, name, publisher, author])
        wb.save("book.xlsx")

        print(str(num) + " - " + id + " - " + name + " - " + publisher + " - " + author)
        num += 1
        toast(
            content="录入成功！",
            duration=1,
            position="center",
            color="info"
        )
        clear_input()


def clear_input():
    """
    清空输入框
    :return: None
    """
    pin.pin["id"] = ""
    pin.pin["isbn"] = ""


css_text = """
body {
    font-family: --apple-system, "华文中宋";
}
"""

# 使用JavaScript修改footer内容
js_text = """
document.getElementsByClassName("footer")[0].innerHTML = "Designed by <a href='https://jike.city/gvenusleo' target='_blank'>GVenusLeo</a>";
"""


@config(
    title="图书信息录入",
    description="扫码获取图书信息并写入Excel表格",
    css_style=css_text,
    js_code=js_text
)
@use_scope("body")
def app():
    """
    主函数
    :return: None
    """
    put_column(
        content=[
            put_html("<h1 align='center'>图书信息录入</h1>"),
            put_row(
                content=[
                    put_column(
                        content=[
                            put_html("<h3 align='center'>图书信息</h3>"),
                            put_scope(
                                "book_info",
                                content=[]
                            ),
                        ],
                        size="20% 80%"
                    ),
                    put_html(""),
                    put_column(
                        content=[
                            put_html("<h3>扫描图书条形码或输入ISBN：</h3>"),
                            pin.put_input(
                                name="isbn",
                                placeholder="ISBN",
                                type="text",
                            ),
                            put_html("<h3>输入易书协编号：</h3>"),
                            pin.put_input(
                                name="id",
                                placeholder="ID",
                                type="text",
                            ),
                            put_buttons(
                                ["提交", "重置"],
                                onclick=[in_sheet, clear_input]
                            )
                        ],
                    )
                ],
                size="45% 10% 45%"
            )
        ],
        size="30% 70%"
    )
    global num
    num = 1
    while True:
        changes = pin.pin_wait_change("isbn", timeout=None)
        global isbn
        isbn = changes["value"]
        if len(isbn) == 13:
            global book_info
            book_info = get_book_info(isbn)
            if book_info == "Error":
                toast(
                    content="API请求错误！",
                    duration=2,
                    position="center",
                    color="error"
                )
                clear_input()
            else:
                with use_scope("book_info", clear=True):
                    put_column(
                        content=[
                            put_html("<h4>书名：" + book_info["name"] + "</h4>"),
                            put_html("<h4>出版：" + book_info["publishing"] + "</h4>"),
                            put_html("<h4>作者：" + book_info["author"] + "</h4>"),
                        ]
                    )


if __name__ == "__main__":
    print("-" * 15 + "图书录入 | 程序启动" + "-" * 15)
    start_server(
        app,
        port=PORT,
        host=HOST,
        cdn=CDN,
        debug=DEBUG,
        auto_open_webbrowser=AUTO_OPEN_BROWSER,
        remote_access=REMOTE_ACCESS,
        static_dir="static",
    )
    print("-" * 15 + "图书录入 | 程序推出" + "-" * 15)